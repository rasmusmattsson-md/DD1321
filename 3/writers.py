import openpyxl

def write_species(people, pet_type, filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Name", "Age", "Email", "Mobile", "Pet Name", "Breed"])
    row = 2

    for person in people.values():
        pets = [p for p in person.pets if p["type"] == pet_type]
        for i, pet in enumerate(pets):
            if i == 0:
                ws.append([person.name, person.age, person.email,
                           person.mobile, pet["name"], pet["breed"]])
            else:
                ws.append(["", "", "", "", pet["name"], pet["breed"]])
            row += 1

    # Statistics
    ws[f"A{row+1}"] = "Average age:"
    ws[f"B{row+1}"] = f"=AVERAGE(B2:B{row-1})"

    ws[f"A{row+2}"] = "Median age:"
    ws[f"B{row+2}"] = f"=MEDIAN(B2:B{row-1})"

    ws[f"A{row+3}"] = "Youngest:"
    ws[f"B{row+3}"] = f"=MIN(B2:B{row-1})"

    ws[f"A{row+4}"] = "Oldest:"
    ws[f"B{row+4}"] = f"=MAX(B2:B{row-1})"

    wb.save(filename)
