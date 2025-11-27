import openpyxl
from models import Person

def load_sheet(path):
    return openpyxl.load_workbook(path).active


def read_dvf(sheet):
    people = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        pnr = str(row[2]).replace("-", "")  # Remove dash, keep as string
        people[pnr] = {
            "name": row[1],
            "person_id": pnr,
            "email": row[3],
            "mobile": row[4],
        }
    return people


def read_hok(sheet):
    people = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue

        name, person_id, pet_type, breed, pet_name = row
        pnr = str(person_id).replace("-", "")  # Normalize to 10 digits 

        if pnr not in people:
            people[pnr] = Person(name, pnr)

        people[pnr].add_pet(pet_type, pet_name, breed)

    return people