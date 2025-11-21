import openpyxl

def write_species(personer, djurtyp, filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Namn", "Ålder", "Epost", "Mobil", "Djurnamn", "Ras"])
    row = 2

    for person in personer.values():
        pets = [p for p in person.husdjur if p["typ"] == djurtyp]
        for i, djur in enumerate(pets):
            if i == 0:
                ws.append([person.namn, person.ålder, person.epost,
                           person.mobil, djur["namn"], djur["ras"]])
            else:
                ws.append(["", "", "", "", djur["namn"], djur["ras"]])
            row += 1

    # Statistik
    ws[f"A{row+1}"] = "Medelålder:"
    ws[f"B{row+1}"] = f"=AVERAGE(B2:B{row-1})"

    ws[f"A{row+2}"] = "Medianålder:"
    ws[f"B{row+2}"] = f"=MEDIAN(B2:B{row-1})"

    ws[f"A{row+3}"] = "Yngst:"
    ws[f"B{row+3}"] = f"=MIN(B2:B{row-1})"

    ws[f"A{row+4}"] = "Äldst:"
    ws[f"B{row+4}"] = f"=MAX(B2:B{row-1})"

    wb.save(filename)
