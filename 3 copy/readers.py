import openpyxl
from models import Person

def load_sheet(path):
    return openpyxl.load_workbook(path).active


def read_dvf(sheet):
    personer = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        personnummer = row[2]
        personer[personnummer] = {
            "namn": row[1],
            "personnummer": personnummer,
            "epost": row[3],
            "mobil": row[4],
        }
    return personer


def read_hok(sheet):
    personer = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue

        namn, personnummer, typ, ras, djurnamn = row

        if personnummer not in personer:
            personer[personnummer] = Person(namn, personnummer)

        personer[personnummer].add_pet(typ, djurnamn, ras)

    return personer
